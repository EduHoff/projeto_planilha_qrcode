import csv
import xlrd
import xlwt
import openpyxl
import os
from pathlib import Path
from entities.column_map import column_map


# =====================================================================
# CSV
# =====================================================================

def count_csv_rows(path: str) -> int:
    if not os.path.exists(path):
        return 0

    with open(path, mode='r', newline='') as file:
        reader = csv.reader(file)

        rows = [row for row in reader if any(cell.strip() for cell in row)]
        return len(rows)


def append_to_csv(path: str, data: list) -> int:
    existing = count_csv_rows(path)
    next_id = existing + 1

    with open(path, mode='a', newline='') as file:
        writer = csv.writer(file)
        writer.writerow([next_id] + data)

    return next_id



# =====================================================================
# XLS (formato antigo)
# =====================================================================

def count_xls_rows(path: str) -> int:
    if not os.path.exists(path):
        return 0

    workbook = xlrd.open_workbook(path)
    sheet = workbook.sheet_by_index(0)

    count = 0
    for row_idx in range(sheet.nrows):
        row = sheet.row_values(row_idx)
        if any(str(cell).strip() for cell in row):
            count += 1
    return count


def append_to_xls(path: str, data: list) -> int:
    if not os.path.exists(path):
        workbook = xlwt.Workbook()
        sheet = workbook.add_sheet("Sheet1")

        next_id = 1
        sheet.write(0, 0, next_id)
        for i, value in enumerate(data, start=1):
            sheet.write(0, i, value)

        workbook.save(path)
        return next_id

    rb = xlrd.open_workbook(path, formatting_info=True)
    r_sheet = rb.sheet_by_index(0)

    existing_rows = count_xls_rows(path)
    next_id = existing_rows + 1

    wb = xlwt.Workbook()
    sheet = wb.add_sheet("Sheet1")

    for row_idx in range(r_sheet.nrows):
        for col_idx in range(r_sheet.ncols):
            sheet.write(row_idx, col_idx, r_sheet.cell_value(row_idx, col_idx))

    new_row = existing_rows
    sheet.write(new_row, 0, next_id)
    for i, value in enumerate(data, start=1):
        sheet.write(new_row, i, value)

    wb.save(path)
    return next_id



# =====================================================================
# XLSX / XLSM (openpyxl)
# =====================================================================

def count_xlsx_rows(path: str) -> int:
    if not os.path.exists(path):
        return 0

    wb = openpyxl.load_workbook(path)
    ws = wb.active

    count = 0
    for row in ws.iter_rows(values_only=True):
        if any(str(cell).strip() for cell in row if cell is not None):
            count += 1
    return count


def append_to_xlsx(path: str, data: list) -> int:
    if not os.path.exists(path):

        wb = openpyxl.Workbook()
        ws = wb.active

        next_id = 1
        ws.append([next_id] + data)

        wb.save(path)
        return next_id

    wb = openpyxl.load_workbook(path)
    ws = wb.active

    existing_rows = count_xlsx_rows(path)
    next_id = existing_rows + 1

    ws.append([next_id] + data)
    wb.save(path)

    return next_id



# =====================================================================
# DISPATCHER GERAL
# =====================================================================


def add_row(path: str | Path, data: list) -> int:

    extension = str(path).lower()

    if extension.endswith(".csv"):
        return append_to_csv(path, data)

    if extension.endswith(".xls"):
        return append_to_xls(path, data)

    if extension.endswith(".xlsx") or extension.endswith(".xlsm"):
        return append_to_xlsx(path, data)

    raise ValueError(f"Formato de arquivo não suportado: {path}")



# ------------------------------------------------------------
# Leitura de CSV sem cabeçalho
# ------------------------------------------------------------
def read_csv_no_header(path: str | Path) -> list[list]:
    if not os.path.exists(path):
        return []

    rows = []
    with open(path, "r", newline="") as file:
        reader = csv.reader(file)
        for row in reader:
            if not any(cell.strip() for cell in row):
                continue
            rows.append(row)

    return rows


# ------------------------------------------------------------
# Filtro
# ------------------------------------------------------------
def filter_rows(path: Path, filters: dict) -> list[dict]:
    data = read_csv_no_header(path)

    result = []
    for row in data:
        ok = True
        for key, value in filters.items():
            col_index = column_map.get(key)

            if col_index is None:
                ok = False
                break

            if str(row[col_index]).lower() != str(value).lower():
                ok = False
                break

        if ok:
            result.append(row)

    return result


# ------------------------------------------------------------
# Leitura completa (sem filtro)
# ------------------------------------------------------------
def read_rows(path: Path) -> list[list]:
    return read_csv_no_header(path)


# ------------------------------------------------------------
# Inserção de nova linha
# ------------------------------------------------------------
def add_row(path: Path, values: list):
    new_id = get_next_id(path)

    row = [new_id] + values

    with open(path, "a", newline="") as file:
        writer = csv.writer(file)
        writer.writerow(row)

    return new_id


def get_next_id(path: Path) -> int:
    rows = read_csv_no_header(path)

    if not rows:
        return 1

    last_id = int(rows[-1][0])
    return last_id + 1